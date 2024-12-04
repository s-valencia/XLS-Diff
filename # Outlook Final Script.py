# Outlook Final Script

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Combines all the gereated summaries from each sheet in the final Differece Report
final_sum = []
def combine_sum():
   combined_string = "\n".join(final_sum)
   return combined_string

# Creates sheet name for new file
def sheet_name(filename):
   if filename.endswith(".xlsx"):
      filename = filename[:-5]
   
   words = filename.replace("_", " ").replace("-", " ").split()
   filtered_words = [word for word in words if word.lower() not in {"pge", "report"}]
   combined_phrase = " ".join(filtered_words)
   
   while len(combined_phrase) > 31:
      filtered_words = filtered_words[:-1]
      combined_phrase = " ".join(filtered_words)

   return combined_phrase
        
# Main function to compare two excel files
output_file = "Difference Reports Test.xlsx"    
def main ():
   # Step 1: Take in two different Excel files
   excel_before_path = input('Enter the (before) excel file name: ')
   if not excel_before_path.endswith('.xlsx'):
       excel_before_path += ".xlsx"

   excel_after_path = input('Enter the (after) excel file name: ')
   if not excel_after_path.endswith('.xlsx'):
       excel_after_path += ".xlsx"

   excel_before = pd.read_excel(excel_before_path)
   excel_after = pd.read_excel(excel_after_path)

   ## Step 2: Create a duplicate of “excel_before” and modify headers
   excel_differences_report = excel_before.copy()
   excluded_headers = ["Service Item ID", "Service Item ArcGIS Online URL", "Item Type"]

   ## Step 3: Add a blank column after every header
   new_columns = []
   for i, column in enumerate(excel_differences_report.columns):
      new_columns.append(column)  # Add the original column
      if i == 0 or (i > 0 and column not in excluded_headers):
         new_columns.append(f"{column} (after)")  # Add a blank column header with a unique name

   # Create the resulting DataFrame with updated column structure
   new_df = pd.DataFrame(columns=new_columns)
   for i, column in enumerate(excel_differences_report.columns):
      new_df[column] = excel_differences_report[column]  # Copy original column data
      if i == 0 or (i > 0 and column not in excluded_headers):
         new_df[f"{column} (after)"] = ""  # Add a blank column with a unique name

   excel_differences_report = new_df

   ## Step 4: Create a duplicate of “excel_after” and modify headers
   excel_after2 = excel_after.copy()
   after_headers = {
      col: f"{col} (after)" if col not in excluded_headers else col for col in excel_after2.columns
   }
   excel_after2.rename(columns=after_headers, inplace=True)


   ## Step 5: Merge all data that have matching headers from “excel_after2” to “XLS Differences Report”
   for column in excel_after2.columns:
      if column in excel_differences_report.columns:
         excel_differences_report[column] = excel_after2[column]

   ## Step 6: Move excluded headers to the first three columns
   ordered_columns = excluded_headers + [col for col in excel_differences_report.columns if col not in excluded_headers]
   excel_differences_report = excel_differences_report[ordered_columns]

   ## Step 7: Analyze for differences and highlight in Excel
   excel_differences_report.to_excel(output_file, index=False, engine='openpyxl')
   wb = load_workbook(output_file)
   sheet = wb.active

   # Define a fill style for highlighting
   highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

   # Start checking from columns 4 and 5, then every two columns after that
   col_index = 4

   while col_index <= sheet.max_column:
      col1 = col_index
      col2 = col_index + 1

      # Break the loop if either header is blank
      if sheet.cell(row=1, column=col1).value is None or sheet.cell(row=1, column=col2).value is None:
         break

      for row in range(2, sheet.max_row + 1):  # Assuming the first row is headers
         value1 = sheet.cell(row=row, column=col1).value
         value2 = sheet.cell(row=row, column=col2).value

         if value1 != value2:  # Highlight mismatched cells
               sheet.cell(row=row, column=col1).fill = highlight_fill
               sheet.cell(row=row, column=col2).fill = highlight_fill

      col_index += 2

   ## Step 8: Delete all rows with no differences
   rows_to_delete = []
   no_difference_rows = 0

   for row in range(2, sheet.max_row + 1):
      differences_found = False
      for col in range(4, sheet.max_column + 1, 2):
         value1 = sheet.cell(row=row, column=col).value
         value2 = sheet.cell(row=row, column=col + 1).value
         if value1 != value2:
               differences_found = True
               break
      if not differences_found:
         rows_to_delete.append(row)
         no_difference_rows += 1

   for row in reversed(rows_to_delete):
      sheet.delete_rows(row)

   ## Step 9: Rename the sheet based on user input
   new_sheet_name = sheet_name(excel_after_path)
   sheet.title = new_sheet_name
      
   # Save the updated workbook
   wb.save(output_file)

   # Step 10 test summary
   differences = []
   for col in excel_differences_report.columns:
      if col.endswith(" (after)"):
         before_col = col.replace(" (after)", "")
         diff_count = (excel_differences_report[col] != excel_differences_report[before_col]).sum()
         if diff_count > 0:
            differences.append(f"'{before_col}': {diff_count} differences found.")

   if differences:
      line_1 = "Key differences found in " + new_sheet_name + ":"
      line_2 = "\n".join([f"- {diff}" for diff in differences])
      combined2 = line_1 + "\n" + line_2
      final_sum.append(combined2)
      final_sum.append("\n")
      sum_deletion = "Deleted Row(s) Summary for " + new_sheet_name + ":"
      sum_deletion2 = str(no_difference_rows) + " row(s) with no differences were deleted."
      final_sum.append(sum_deletion)
      final_sum.append(sum_deletion2)
   else:
      combined2 = "No differences were found."
      final_sum.append(combined2)
      final_sum.append("\n")
      sum_deletion = "Deleted Row(s) Summary for " + new_sheet_name + ": "
      sum_deletion2 = str(no_difference_rows) + " row(s) with no differences were deleted."
      final_sum.append(sum_deletion)
      final_sum.append(sum_deletion2)

# Run the main function
main()

# Reruns the script using two new excel files
def rerun(new_file):
   # Step 0: Take in two different Excel files
    xlsx_before_path = input('Enter the next (before) excel filename: ')
    if not xlsx_before_path.endswith('.xlsx'):
        xlsx_before_path += ".xlsx"

    xlsx_after_path = input('Enter the next (after) excel filename: ')
    if not xlsx_after_path.endswith('.xlsx'):
        xlsx_after_path += ".xlsx"
   
    excel_before = pd.read_excel(xlsx_before_path)
    excel_after = pd.read_excel(xlsx_after_path)
    
    ## Step 1: create new tab in the excel file
    wb = load_workbook(new_file)
    new_sheet_name = sheet_name(xlsx_after_path)
    wb.create_sheet(title=new_sheet_name)

    ## Step 2: Create a duplicate of “excel_before” and modify headers
    excel_report = excel_before.copy()
    exclude_headers = ["Service Item ID", "Service Item ArcGIS Online URL", "Item Type"]
    
    ## Step 3: Add a blank column after every header in before excel
    new_columns = []
    for i, column in enumerate(excel_report.columns):
        new_columns.append(column)  # Add the original column
        if i == 0 or (i > 0 and column not in exclude_headers):
            new_columns.append(f"{column} (after)")  # Add a blank column header with a unique name
    
    # Create the resulting DataFrame with updated column structure
    new_df = pd.DataFrame(columns=new_columns)
    for i, column in enumerate(excel_report.columns):
        new_df[column] = excel_report[column]  # Copy original column data
        if i == 0 or (i > 0 and column not in exclude_headers):
            new_df[f"{column} (after)"] = ""  # Add a blank column with a unique name
    
    excel_report = new_df

    ## Step 4: Create a duplicate of “excel_after” and modify headers
    excel_after2 = excel_after.copy()
    after_headers = {
        col: f"{col} (after)" if col not in exclude_headers else col for col in excel_after2.columns
    }
    excel_after2.rename(columns=after_headers, inplace=True)

    ## Step 5: Merge all data that have matching headers from “excel_after2” to “XLS Differences Report”
    for column in excel_after2.columns:
        if column in excel_report.columns:
            excel_report[column] = excel_after2[column]
        
    ## Step 6: Move excluded headers to the far left of all columns
    valid_exclude_headers = [header for header in exclude_headers if header in excel_report.columns]
    ordered_columns = valid_exclude_headers + [col for col in excel_report.columns if col not in valid_exclude_headers]
    excel_report = excel_report[ordered_columns]

    with pd.ExcelWriter(new_file, engine='openpyxl', mode='a') as writer:
        excel_report.to_excel(writer, sheet_name=new_sheet_name, index=False)
    
    ## Step 7: Analyze for differences and highlight in Excel

    # Define a fill style for highlighting
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Reload the workbook and new sheet to ensure data is written
    wb = load_workbook(new_file)
    new_sheet = wb[new_sheet_name]

    # Iterate through columns to find headers with " (after)"
    for col in range(1, new_sheet.max_column + 1):
        header = new_sheet.cell(row=1, column=col).value
        if header and "after" in header:
            before_col = col - 1  # Column to the left of the "after" column

            for row in range(2, new_sheet.max_row + 1):  # Assuming the first row is headers
                value1 = new_sheet.cell(row=row, column=before_col).value
                value2 = new_sheet.cell(row=row, column=col).value

                if value1 != value2:  # Highlight mismatched cells
                    new_sheet.cell(row=row, column=before_col).fill = highlight_fill
                    new_sheet.cell(row=row, column=col).fill = highlight_fill

    wb.save(new_file)

    
    ## Step 8: Delete all rows with no highlighted cells
    rows_to_delete = []
    no_difference_rows = 0

    for row in range(2, new_sheet.max_row + 1):
        highlighted_cells_found = False
        for col in range(1, new_sheet.max_column + 1):
            cell = new_sheet.cell(row=row, column=col)
            if cell.fill == highlight_fill:
                highlighted_cells_found = True
                break
        if not highlighted_cells_found:
            rows_to_delete.append(row)
            no_difference_rows += 1

    for row in reversed(rows_to_delete):
        new_sheet.delete_rows(row)
    
    wb.save(new_file)

    # Step 9: Summary test
    differences2 = []
    for col in excel_report.columns:
        if col.endswith(" (after)"):
            before_col = col.replace(" (after)", "")
            diff_count = (excel_report[col] != excel_report[before_col]).sum()
            if diff_count > 0:
                differences2.append(f"'{before_col}': {diff_count} differences found.")

    if differences2:
        line_1 = "Key differences found in " + new_sheet_name + ":"
        line_2 = "\n".join([f"- {diff}" for diff in differences2])
        combined2 = line_1 + "\n" + line_2
        final_sum.append("\n")
        final_sum.append(combined2)
        final_sum.append("\n")
        sum_deletion = "Deleted Row(s) Summary for " + new_sheet_name + ":"
        sum_deletion2 = str(no_difference_rows) + " row(s) with no differences were deleted."
        final_sum.append(sum_deletion)
        final_sum.append(sum_deletion2)
    else:
        combined2 = "No differences were found."
        final_sum.append("\n")
        final_sum.append(combined2)
        final_sum.append("\n")
        sum_deletion = "Deleted Row(s) Summary for " + new_sheet_name + ": "
        sum_deletion2 = str(no_difference_rows) + " row(s) with no differences were deleted."
        final_sum.append(sum_deletion)
        final_sum.append(sum_deletion2)

# Send email with the difference report
# The summary of the differences is embedded in body of the email
def send_email():
    msg = MIMEMultipart()
    msg.attach(MIMEText(combine_sum(), 'plain'))
    msg['Subject'] = "Difference Report: XLS"
    msg['From'] = 'donotreply@pge.com'
    msg['To'] = 'jk56@pge.com'

    # Output file attachment
    with open(output_file, 'rb') as f:
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(f.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=output_file)
        msg.attach(attachment)

    try:
        with smtplib.SMTP('mailhost', 25) as server:
            server.sendmail(msg['From'], [msg['To']], msg.as_string())
        print("Email sent successfully.")
    except Exception as e:
        print(f"Failed to send email: {e}")
        
while True:
    rerun_choice = input("Would you like to compare two new excel files? ").lower()
    if rerun_choice in ['y', 'yes']:
        rerun(output_file)
    else:
       send_email()
       break