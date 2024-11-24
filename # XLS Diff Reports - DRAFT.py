# XLS Diff Reports - DRAFT
# Description: This script is designed to take two Excel files as input, 
#              create a duplicate of the first file with updated headers, 
#              highlight differences between the two files, 
#              store it in a sheet and be able to repeat the process several times, 
#              and send an email with the xlsx differences file as an attachment with a summary.

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    ## Step 1: Take in two different Excel files
    excel_before_path = 'PGE_Report_Field_Maps before - Copy.xlsx'
    excel_after_path = 'PGE_Report_Field_Maps after - Copy.xlsx'
    # excel_before_path = input('Enter the first excel file: ')
    # excel_after_path = input('Enter the second excel file: ')

    excel_before = pd.read_excel(excel_before_path)
    excel_after = pd.read_excel(excel_after_path)

    ## Step 2: Create a duplicate of “excel_before” and modify headers
    excel_differences_report = excel_before.copy()
    excluded_headers = ["Service Item ID", "Service Item ArcGIS Online URL", "Item Type"]
    # excluded_headers = input("Enter the headers to exclude when comparing excel files (comma-separated): ").split(", ")

    ## Step 3: Add a blank column after every header
    new_columns = []
    for i, column in enumerate(excel_differences_report.columns):
        new_columns.append(column)  # Add the original column
        if i == 0 or (i > 0 and column not in excluded_headers):
            new_columns.append(f"{column} (after)")  # Add a blank column header with a unique name
    
    # Create the resulting DataFrame with updated column structure
    new_df = pd.DataFrame(columns=new_columns)
    for i, column in enumerate(excel_differences_report.columns):
        new_df[column] = excel_differences_report[column]  # Copy original column data
        if i == 0 or (i > 0 and column not in excluded_headers):
            new_df[f"{column} (after)"] = ""  # Add a blank column with a unique name
    
    excel_differences_report = new_df

    ## Step 4: Create a duplicate of “excel_after” and modify headers
    excel_after2 = excel_after.copy()
    after_headers = {
        col: f"{col} (after)" if col not in excluded_headers else col for col in excel_after2.columns
    }
    excel_after2.rename(columns=after_headers, inplace=True)


    ## Step 5: Merge all data that have matching headers from “excel_after2” to “XLS Differences Report”
    for column in excel_after2.columns:
        if column in excel_differences_report.columns:
            excel_differences_report[column] = excel_after2[column]

    ## Step 6: Move excluded headers to the first three columns
    ordered_columns = excluded_headers + [col for col in excel_differences_report.columns if col not in excluded_headers]
    excel_differences_report = excel_differences_report[ordered_columns]
    
    ## Step 7: Analyze for differences and highlight in Excel
    output_file = input("Enter the output file name: ") + ".xlsx"
    excel_differences_report.to_excel(output_file, index=False, engine='openpyxl')
    wb = load_workbook(output_file)
    sheet = wb.active

    # Define a fill style for highlighting
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Start checking from columns 4 and 5, then every two columns after that
    col_index = 4

    while col_index <= sheet.max_column:
        col1 = col_index
        col2 = col_index + 1

        # Break the loop if either header is blank
        if sheet.cell(row=1, column=col1).value is None or sheet.cell(row=1, column=col2).value is None:
            break

        for row in range(2, sheet.max_row + 1):  # Assuming the first row is headers
            value1 = sheet.cell(row=row, column=col1).value
            value2 = sheet.cell(row=row, column=col2).value

            if value1 != value2:  # Highlight mismatched cells
                sheet.cell(row=row, column=col1).fill = highlight_fill
                sheet.cell(row=row, column=col2).fill = highlight_fill

        col_index += 2
    
    ## Step 8: Delete all rows with no differences
    rows_to_delete = []
    no_difference_rows = 0

    for row in range(2, sheet.max_row + 1):
        differences_found = False
        for col in range(4, sheet.max_column + 1, 2):
            value1 = sheet.cell(row=row, column=col).value
            value2 = sheet.cell(row=row, column=col + 1).value
            if value1 != value2:
                differences_found = True
                break
        if not differences_found:
            rows_to_delete.append(row)
            no_difference_rows += 1

    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)

    ## Step 9: Rename the sheet based on user input
    new_sheet_name = input("Enter the new sheet name: ")
    sheet.title = new_sheet_name
        
    # Save the updated workbook
    wb.save(output_file)
    
    ## Step 10: Print a brief summary and bullet points of differences found
    differences = []
    for col in excel_differences_report.columns:
        if col.endswith(" (after)"):
            before_col = col.replace(" (after)", "")
            diff_count = (excel_differences_report[col] != excel_differences_report[before_col]).sum()
            if diff_count > 0:
                differences.append(f"'{before_col}': {diff_count} differences found.")

    if differences:
        print(f"Summary of Differences for '{new_sheet_name}': ")
        print("\n".join([f"- {diff}" for diff in differences]))
    else:
        print(f"No differences were found between the two Excel files for '{new_sheet_name}'.")

    print(f"\nSummary of Deleted Rows for '{new_sheet_name}': ")
    print(f"- {no_difference_rows} rows with no differences were deleted.")

if __name__ == "__main__":
    main()